"""
Local API server — reads ARE_BID_TRACKER_2026.xlsx and serves JSON to the React app.
Run: source venv/bin/activate && python api.py
"""

from flask import Flask, jsonify, request
from flask_cors import CORS
from pathlib import Path
import hashlib
import sys
import uuid
import openpyxl
import psycopg2
import psycopg2.extras

app = Flask(__name__)
# M-17 (2026-07-23): this Flask API has no authentication, so it must not be
# openly reachable. CORS is restricted to the local Bid Tracker UI origin (the
# browser only ever calls the API through the vite proxy at same-origin :5173,
# so this does not affect normal use — it blocks arbitrary websites from
# scripting the unauthenticated API via the user's browser). Genuine auth is
# tracked as M-18.
CORS(app, origins=["http://localhost:5173", "http://127.0.0.1:5173"])

AGENT_DIR = Path(
    "/Users/andyramos/Developer"
    "/06_Pyhton_Scripts/02_Bid_Pyhton_Scripts/ARE_BID_TRACKER_AGENT"
)
EXCEL_FILE  = AGENT_DIR / "ARE_BID_TRACKER_2026.xlsx"
SAM_RAW_DIR = AGENT_DIR / "sam_raw"

# Project Master cutover (2026-07-08, owner-approved): /api/projects reads and
# writes Postgres now, not the Excel "Project " sheet -- same Postgres
# `project` table Financial Hub's main.py uses, connected to directly since
# this Flask process has no ORM/session layer of its own. /api/bids,
# /api/recompete, and /api/health are unaffected -- still Excel-backed, out of
# scope for this cutover (Project Master only).
def _pg_dsn() -> str:
    env_path = Path(
        "/Users/andyramos/Developer/06_Pyhton_Scripts"
        "/01_Finance_Phyton_Scripts/ARE_FINANCIAL_HUB/backend/.env"
    )
    for line in env_path.read_text().splitlines():
        k, _, v = line.partition("=")
        if k.strip() == "DATABASE_URL":
            return v.strip()
    raise RuntimeError("DATABASE_URL not found in Financial Hub's .env")


def _pg_conn():
    return psycopg2.connect(_pg_dsn(), cursor_factory=psycopg2.extras.RealDictCursor)

# Same cross-process lock + atomic save Financial Hub's backend/utils/file_locking.py
# and ARE_BID_TRACKER_AGENT's other writers use against this same workbook — without
# it, this process's writes race unlocked against theirs (reproducible zipfile.BadZipFile
# corruption). Reused, not reimplemented, so the lock path derivation can't drift.
sys.path.insert(0, str(AGENT_DIR))
from excel_lock_utils import workbook_lock, atomic_save  # noqa: E402


def _build_sam_lookup():
    """Build {notice_id: {due_date, agency}, title_lower: {due_date, agency}} from raw SAM JSON."""
    lookup_id    = {}
    lookup_title = {}
    if not SAM_RAW_DIR.exists():
        return lookup_id, lookup_title
    import json
    for f in sorted(SAM_RAW_DIR.glob("*.json")):
        try:
            data = json.loads(f.read_text())
        except Exception:
            continue
        for op in data.get("opportunitiesData", []):
            due = (op.get("responseDeadLine") or op.get("archiveDate") or "")[:10]
            full_path = op.get("fullParentPathName") or ""
            agency = full_path.split(".")[-1].strip() if full_path else ""
            notice_id = (op.get("noticeId") or "").lower()
            title_key = (op.get("title") or "").strip().lower()
            entry = {"due_date": due, "agency": agency}
            if notice_id:
                lookup_id[notice_id] = entry
            if title_key:
                lookup_title[title_key] = entry
    return lookup_id, lookup_title


def _fmt_date(val):
    """Convert Excel datetime objects or strings to YYYY-MM-DD."""
    if not val:
        return ""
    from datetime import datetime as dt
    if isinstance(val, dt):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    # Handle "4/13/26 0:00" style
    for fmt in ("%m/%d/%y %H:%M", "%m/%d/%Y %H:%M", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return dt.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return s[:10]  # fallback: first 10 chars


def _score_to_priority(score):
    try:
        s = int(score or 0)
    except (TypeError, ValueError):
        s = 0
    if s >= 100: return "Critical"
    if s >= 10:  return "High"
    if s >= 5:   return "Medium"
    return "Low"


def _load_sheet(tab_name):
    if not EXCEL_FILE.exists():
        return None, []
    wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True, data_only=True)
    if tab_name not in wb.sheetnames:
        return None, []
    ws = wb[tab_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return None, []
    headers = [str(h).strip().replace('\xa0', '').strip() if h else "" for h in rows[0]]
    return headers, rows[1:]


def _latest_hits_tab():
    if not EXCEL_FILE.exists():
        return None
    wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True, data_only=True)
    for sheet in sorted(wb.sheetnames, reverse=True):
        if sheet.startswith("Hits "):
            return sheet
    return None


@app.route("/api/bids")
def get_bids():
    # Serve from the Bid Tracker tab — these are the filtered strong leads
    headers, data_rows = _load_sheet("Bid Tracker")
    if not headers:
        return jsonify([])

    import re
    sam_by_id, sam_by_title = _build_sam_lookup()

    bids = []
    for i, row in enumerate(data_rows):
        d = dict(zip(headers, row))
        if not d.get("Bid / Opportunity Name"):
            continue

        score = d.get("AI Score") or 0
        try:
            score = int(score)
        except (TypeError, ValueError):
            score = 0

        status_raw = str(d.get("Status") or "Identified")
        status_map = {"Identified": "Open", "Submitted": "Open", "Won": "Awarded",
                      "Lost": "Closed", "Archived": "Closed",
                      # Capture-engine decisions are active leads:
                      "Recommended": "Open", "Manual Review": "Open",
                      "Prospect": "Open", "Active Bid": "Open",
                      "Submitted": "Open", "Awarded": "Awarded",
                      "Closed Won": "Closed", "Closed Lost": "Closed",
                      "Closed Cancelled": "Closed", "Closed Withdrawn": "Closed"}
        status = status_map.get(status_raw, "Open")

        # Capture Intelligence decision fields (blank on legacy rows).
        def _int(v):
            try: return int(float(v))
            except (TypeError, ValueError): return 0
        def _float(v):
            try: return round(float(v), 3)
            except (TypeError, ValueError): return 0.0
        decision = str(d.get("Decision") or "").strip()

        notes = str(d.get("Notes / Rationale") or "")
        # SAM Link: dedicated column first, fall back to regex from Notes
        link = str(d.get("SAM Link") or "").strip()
        if not link:
            m = re.search(r"https?://\S+", notes)
            link = m.group(0) if m else ""

        # Due date and agency: use Excel value, fall back to SAM raw lookup
        notice_id = str(d.get("Solicitation / Notice ID") or "").lower()
        title_key = str(d.get("Bid / Opportunity Name") or "").strip().lower()
        sam_info  = sam_by_id.get(notice_id) or sam_by_title.get(title_key) or {}

        due    = _fmt_date(d.get("Due Date")) or sam_info.get("due_date", "")
        agency = str(d.get("Agency / Department") or "") or sam_info.get("agency", "")
        # Ignore placeholder values left from old data
        if agency in ("Prompt", "VA Medical Center") and sam_info.get("agency"):
            agency = sam_info["agency"]

        prefix = _parse_prefix(str(d.get("Tags Prefix") or ""))

        bids.append({
            "id":            str(d.get("Solicitation / Notice ID") or f"bt-{i}"),
            "status":        status,
            "workflowStatus": status_raw,   # raw pipeline stage (Recommended/Manual Review/Prospect/Active Bid/...)
            "dueDate":       due,
            "title":         str(d.get("Bid / Opportunity Name") or ""),
            "state":         str(d.get("State") or ""),
            "city":          str(d.get("City")  or ""),
            "facility":      agency or "VA Medical Center",
            "bidAmount":     str(d.get("Bid Amount") or ""),
            "awardedAmount": str(d.get("Award Amount") or ""),
            "contractor":    "",
            "priority":      _score_to_priority(score),
            "category":      prefix["category"] or "Electrical",
            "starred":       prefix["starred"],
            "link":          link,
            "score":         score,
            "whyItScored":   notes,
            "setAside":      "",
            "triennial":     "triennial" in notes.lower(),
            "maintenance":   "maintenance" in notes.lower(),
            "postedDate":    str(d.get("Date Added") or ""),
            "naics":         str(d.get("NAICS") or ""),
            "folderPath":    str(d.get("Folder Path") or ""),
            # ── Capture Intelligence decision fields ──
            "decision":             decision or ("Recommended" if score >= 85 else ""),
            "finalScore":           _int(d.get("Final Score") or score),
            "decisionReason":       str(d.get("Decision Reason") or ""),
            "capabilityCount":      _int(d.get("Capability Count")),
            "blacklistHits":        _int(d.get("Blacklist Hits")),
            "historicalSimilarity": _float(d.get("Historical Similarity")),
            "confidence":           _int(d.get("Confidence")),
            "lastReviewDate":       str(d.get("Last Review Date") or ""),
            # ── Phase 3B promotion fields ──
            "financialHubProjectId": str(d.get("Financial Hub Project ID") or ""),
            "jobNumber":             str(d.get("Job Number") or ""),
            "promotionStatus":       str(d.get("Promotion Status") or ""),
            "promotionDate":         str(d.get("Promotion Date") or ""),
            # ── Phase 3G lifecycle fields ──
            "submittedDate":         str(d.get("Submitted Date") or ""),
            "submittedBy":           str(d.get("Submitted By") or ""),
            "submissionStatus":      str(d.get("Submission Status") or ""),
            "closedDate":            str(d.get("Closed Date") or ""),
            "closedReason":          str(d.get("Closed Reason") or ""),
            "resultStatus":          str(d.get("Result Status") or ""),
            # ── Phase 3H command-center meta ──
            "awardProbability":      _int(d.get("Award Probability")),
            "followUpStatus":        str(d.get("Follow-Up Status") or ""),
            "lastFollowUp":          str(d.get("Last Follow-Up Date") or ""),
            "nextFollowUp":          str(d.get("Next Follow-Up Date") or ""),
            "followUpNotes":         str(d.get("Follow-Up Notes") or ""),
            "coName":                str(d.get("CO Name") or ""),
            "coEmail":               str(d.get("CO Email") or ""),
            "awardDecisionDate":     str(d.get("Award Decision Date") or ""),
            "debriefRequested":      str(d.get("Debrief Requested") or ""),
            "debriefReceived":       str(d.get("Debrief Received") or ""),
            "notes":         [],
            "chk_sf1449":     bool(d.get("SF1449")),
            "chk_sow_pws":    bool(d.get("SOW/PWS")),
            "chk_pricing":    bool(d.get("Pricing")),
            "chk_past_perf":  bool(d.get("Past Perf")),
            "chk_osha_safety": bool(d.get("OSHA")),
            "chk_licenses":   bool(d.get("Licenses")),
            "chk_site_visit": bool(d.get("Site Visit")),
            "chk_sub_loi":    bool(d.get("Sub LOI")),
            "chk_compliance": bool(d.get("Compliance")),
        })

    return jsonify(bids)


@app.route("/api/recompete")
def get_recompete():
    headers, data_rows = _load_sheet("Recompete Watch")
    if not headers:
        return jsonify([])

    rows = []
    for row in data_rows:
        d = dict(zip(headers, row))
        rows.append({
            "status":        str(d.get("Status") or ""),
            "daysUntil":     d.get("Days Until") or 0,
            "city":          str(d.get("City")   or ""),
            "state":         str(d.get("State")  or ""),
            "recompeteDate": str(d.get("Recompete Date") or ""),
            "awardDate":     str(d.get("Award Date") or ""),
            "cycle":         str(d.get("Cycle")  or ""),
            "contractor":    str(d.get("Contractor") or ""),
            "areWin":        str(d.get("ARE Win?") or "NO"),
            "amount":        str(d.get("Est. Amount") or ""),
            "contractId":    str(d.get("Contract ID") or ""),
            "title":         str(d.get("Title") or ""),
        })

    return jsonify(rows)


PROJECTS_JSON = AGENT_DIR / "exports" / "projects.json"


def _read_pg_projects():
    """Postgres-backed replacement for the old Excel "Project " sheet reader.
    Same response shape the frontend already expects — milestones/invoices are
    still synthesized from flat fields (milestones_text, invoiced/invoiced2/
    pending_paid) exactly as the Excel version did (there was never a
    structured milestones table behind this particular view). issues/notes
    now come from the real project_issue/project_note tables Financial Hub's
    project_ops_router already writes to — a genuine improvement over the
    Excel version, which only ever populated notes and left issues empty."""
    conn = _pg_conn()
    try:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT excel_project_id, project_name, category, customer_name, status, phase,
                       start_date, end_date, contract_amount, collected, sub_value, material,
                       profit, milestones_text, invoiced, invoiced2, pending_paid, onedrive_path
                FROM project
                WHERE is_archived = false AND excel_project_id IS NOT NULL
            """)
            rows = cur.fetchall()

            projects, project_by_id = [], {}
            for r in rows:
                pid = r["excel_project_id"]
                contract = float(r["contract_amount"] or 0)
                collected = float(r["collected"] or 0) or float(r["invoiced"] or 0) + float(r["invoiced2"] or 0)
                milestones_raw = r["milestones_text"] or ""
                milestones = [{"id": f"m-{j}", "title": m.strip(), "completed": False}
                              for j, m in enumerate(milestones_raw.split(",")) if m.strip()]
                invoices = []
                if r["invoiced"]:      invoices.append({"id": "inv-1", "amount": float(r["invoiced"]),      "status": "Paid"})
                if r["invoiced2"]:     invoices.append({"id": "inv-2", "amount": float(r["invoiced2"]),     "status": "Paid"})
                if r["pending_paid"]:  invoices.append({"id": "inv-p", "amount": float(r["pending_paid"]),  "status": "Pending"})

                proj = {
                    "id":             pid,
                    "title":          r["project_name"],
                    "category":       r["category"] or "",
                    "facility":       r["customer_name"] or "",
                    "status":         r["status"] or "In Progress",
                    "phase":          r["phase"] or "Execution",
                    "progress":       int(min(100, (collected / contract * 100))) if contract else 0,
                    "startDate":      r["start_date"].isoformat() if r["start_date"] else "",
                    "endDate":        r["end_date"].isoformat() if r["end_date"] else "",
                    "contractValue":  contract,
                    "collectedValue": collected,
                    "subContractor":  float(r["sub_value"] or 0),
                    "material":       float(r["material"] or 0),
                    "profit":         float(r["profit"] or 0),
                    "milestones":     milestones,
                    "invoices":       invoices,
                    "issues":         [],
                    "notes":          [],
                    "onedriveFolder": r["onedrive_path"] or "",
                }
                projects.append(proj)
                project_by_id[pid] = proj

            if project_by_id:
                cur.execute(
                    "SELECT project_id, title, status FROM project_issue WHERE is_deleted = false AND project_id = ANY(%s)",
                    (list(project_by_id.keys()),),
                )
                for row in cur.fetchall():
                    project_by_id[row["project_id"]]["issues"].append(
                        {"id": row["title"], "title": row["title"], "status": row["status"]}
                    )
                cur.execute(
                    "SELECT project_id, text FROM project_note WHERE is_deleted = false AND project_id = ANY(%s)",
                    (list(project_by_id.keys()),),
                )
                for row in cur.fetchall():
                    project_by_id[row["project_id"]]["notes"].append(row["text"])

            return projects
    finally:
        conn.close()


@app.route("/api/projects", methods=["GET"])
def get_projects():
    return jsonify(_read_pg_projects())


@app.route("/api/projects", methods=["POST"])
def save_project():
    """Upsert a project into Postgres. Milestones stay a flat comma-joined
    string (milestones_text) matching the pre-cutover Excel behavior exactly.
    Issues/notes upsert into the real project_issue/project_note tables via a
    delete-then-reinsert for this project_id, mirroring the old
    _upsert_normalized_tab's semantics on the tables that now matter."""
    proj = request.get_json(force=True)

    # Stable ID: reuse the one the client sent back (from a prior GET) if present,
    # otherwise this is a brand-new project — mint one now, once, and never derive
    # it from mutable fields like title again.
    proj_id = str(proj.get("id") or "").strip() or f"proj-{uuid.uuid4().hex[:8]}"

    milestones = proj.get("milestones") or []
    invoices   = proj.get("invoices")   or []
    issues     = proj.get("issues")     or []
    notes      = proj.get("notes")      or []

    paid_total    = sum(i.get("amount", 0) for i in invoices if i.get("status") == "Paid")
    pending_total = sum(i.get("amount", 0) for i in invoices if i.get("status") == "Pending")
    milestone_str = ", ".join(m.get("title", "") for m in milestones if m.get("title"))

    conn = _pg_conn()
    try:
        with conn.cursor() as cur:
            cur.execute("SELECT id FROM project WHERE excel_project_id = %s", (proj_id,))
            existing = cur.fetchone()

            if existing:
                cur.execute("""
                    UPDATE project SET
                        project_name = %s, category = %s, customer_name = %s, status = %s, phase = %s,
                        start_date = %s, end_date = %s, contract_amount = %s, collected = %s,
                        onedrive_path = %s, milestones_text = %s, invoiced = %s, pending_paid = %s,
                        updated_at = now()
                    WHERE excel_project_id = %s
                """, (
                    proj.get("title", ""), proj.get("category"), proj.get("facility", ""),
                    proj.get("status", "In Progress"), proj.get("phase", "Planning"),
                    proj.get("startDate") or None, proj.get("endDate") or None,
                    proj.get("contractValue") or 0, paid_total or proj.get("collectedValue") or 0,
                    proj.get("onedriveFolder"), milestone_str, paid_total or None, pending_total or None,
                    proj_id,
                ))
            else:
                cur.execute("SELECT to_char(now(), 'YY') AS yy")
                yy = cur.fetchone()["yy"]
                cur.execute(
                    "INSERT INTO job_number_sequence (year, last_seq) VALUES (%s, 1) "
                    "ON CONFLICT (year) DO UPDATE SET last_seq = job_number_sequence.last_seq + 1 "
                    "RETURNING last_seq",
                    (2000 + int(yy),),
                )
                seq = cur.fetchone()["last_seq"]
                job_number = f"ARE-{yy}-{seq:03d}"
                cur.execute("""
                    INSERT INTO project (
                        job_number, project_name, customer_name, category, status, phase,
                        start_date, end_date, contract_amount, revised_contract_amount,
                        estimated_cost, estimated_margin, collected, onedrive_path,
                        milestones_text, invoiced, pending_paid, excel_project_id, origin,
                        created_by, is_archived, created_at, updated_at
                    ) VALUES (
                        %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 0, 0, %s, %s, %s, %s, %s, %s,
                        'manual', 'bid_tracker', false, now(), now()
                    )
                """, (
                    job_number, proj.get("title", ""), proj.get("facility", ""), proj.get("category"),
                    proj.get("status", "In Progress"), proj.get("phase", "Planning"),
                    proj.get("startDate") or None, proj.get("endDate") or None,
                    proj.get("contractValue") or 0, proj.get("contractValue") or 0,  # revised_contract_amount = contract_amount on create, no COs yet
                    paid_total or proj.get("collectedValue") or 0,
                    proj.get("onedriveFolder"), milestone_str, paid_total or None, pending_total or None,
                    proj_id,
                ))

            cur.execute("UPDATE project_issue SET is_deleted = true WHERE project_id = %s", (proj_id,))
            for iss in issues:
                cur.execute(
                    "INSERT INTO project_issue (project_id, title, priority, status, source, created_by, created_at, is_deleted) "
                    "VALUES (%s, %s, %s, %s, 'bid_tracker', 'bid_tracker', now(), false)",
                    (proj_id, iss.get("title", ""), iss.get("priority", "Medium"), iss.get("status", "Open")),
                )

            cur.execute("UPDATE project_note SET is_deleted = true WHERE project_id = %s", (proj_id,))
            for n in notes:
                cur.execute(
                    "INSERT INTO project_note (project_id, text, source, created_by, created_at, is_deleted) "
                    "VALUES (%s, %s, 'bid_tracker', 'bid_tracker', now(), false)",
                    (proj_id, n),
                )

        conn.commit()
    finally:
        conn.close()

    return jsonify({"ok": True, "id": proj_id})


@app.route("/api/open-folder")
def open_folder():
    import subprocess
    path = request.args.get("path", "").strip()
    if not path:
        return jsonify({"error": "No path provided"}), 400
    # Security: only allow paths inside OneDrive project folder
    allowed = (
        "/Users/andyramos/Developer/03_Project",
        "/Users/andyramos/Developer/02_Bid_Estimating",
    )
    if not any(path.startswith(p) for p in allowed):
        return jsonify({"error": "Path not allowed"}), 403
    subprocess.Popen(["open", path])
    return jsonify({"ok": True})


# Map app chk_* keys → spreadsheet column headers (U:AC)
_CHK_COLS = {
    "chk_sf1449":     "SF1449",
    "chk_sow_pws":    "SOW/PWS",
    "chk_pricing":    "Pricing",
    "chk_past_perf":  "Past Perf",
    "chk_osha_safety": "OSHA",
    "chk_licenses":   "Licenses",
    "chk_site_visit": "Site Visit",
    "chk_sub_loi":    "Sub LOI",
    "chk_compliance": "Bid Submitted",
}

def _parse_prefix(s: str) -> dict:
    """Parse 'category:X;priority:Y;starred:false' from the Tags Prefix column (AD)."""
    out = {"category": "Electrical", "priority": "", "starred": False}
    for chunk in (s or "").split(";"):
        if ":" not in chunk:
            continue
        k, v = chunk.split(":", 1)
        k, v = k.strip(), v.strip()
        if k == "starred":
            out["starred"] = v.lower() == "true"
        elif k in ("category", "priority"):
            out[k] = v
    return out

def _build_prefix_string(bid: dict) -> str:
    """Build the Tags Prefix string for column AD (category/priority/starred only)."""
    return (
        f"category:{bid.get('category', 'Electrical')};"
        f"priority:{bid.get('priority', 'Medium')};"
        f"starred:{'true' if bid.get('starred') else 'false'}"
    )


@app.route("/api/bids", methods=["POST"])
def save_bid():
    """Upsert a bid into the 'Bid Tracker' sheet by Solicitation / Notice ID."""
    import datetime as _dt
    bid = request.get_json(force=True)
    if not EXCEL_FILE.exists():
        return jsonify({"error": "Excel file not found"}), 500

    with workbook_lock(EXCEL_FILE):
        wb = openpyxl.load_workbook(EXCEL_FILE, data_only=False)
        if "Bid Tracker" not in wb.sheetnames:
            return jsonify({"error": "Sheet 'Bid Tracker' not found"}), 500

        ws = wb["Bid Tracker"]
        raw_headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]

        # Map normalised key → actual column index (1-based)
        def col_idx(name):
            for i, h in enumerate(raw_headers):
                if h.replace('\xa0', '').strip().lower() == name.lower():
                    return i + 1
            return None

        # Field → Excel column lookup (use actual Excel header text)
        # NOTE: do NOT write to "AI Tags" (M) — it is a formula built from U:AC + AD
        FIELD_MAP = {
            "Solicitation / Notice ID": bid.get("id", ""),
            "Status":                   bid.get("status", "Open"),
            "Due Date":                 bid.get("dueDate", ""),
            "Bid / Opportunity Name":   bid.get("title", ""),
            "State":                    bid.get("state", ""),
            "City":                     bid.get("city", ""),
            "Agency / Department":      bid.get("facility", ""),
            "Bid Amount":               bid.get("bidAmount", ""),
            "Award Amount":             bid.get("awardedAmount", ""),
            "SAM Link":                 bid.get("link", ""),
            "Folder Path":              bid.get("folderPath", ""),
            "Notes / Rationale":        "; ".join(bid.get("notes") or []),
            "Tags Prefix":              _build_prefix_string(bid),
            "Last Updated":             _dt.datetime.now().strftime("%Y-%m-%d"),
        }

        # Checklist values written to individual boolean columns U:AC
        checklist_values = {hdr: bool(bid.get(key)) for key, hdr in _CHK_COLS.items()}

        # Find existing row: by Notice ID first, fall back to title for auto-generated bt-N ids
        notice_id = str(bid.get("id", "")).strip()
        is_auto_id = notice_id.startswith("bt-")
        target_row = None

        id_col    = col_idx("Solicitation / Notice ID")
        title_col = col_idx("Bid / Opportunity Name")

        if not is_auto_id and notice_id and id_col:
            for row in ws.iter_rows(min_row=2):
                if str(row[id_col - 1].value or "").strip() == notice_id:
                    target_row = row[0].row
                    break

        if target_row is None and title_col:
            bid_title = str(bid.get("title", "")).strip().lower()
            for row in ws.iter_rows(min_row=2):
                if str(row[title_col - 1].value or "").strip().lower() == bid_title:
                    target_row = row[0].row
                    break

        if target_row is None:
            # Append new row
            new_row = [""] * len(raw_headers)
            for col_name, value in FIELD_MAP.items():
                idx = col_idx(col_name)
                if idx:
                    new_row[idx - 1] = value
            for col_name, value in checklist_values.items():
                idx = col_idx(col_name)
                if idx:
                    new_row[idx - 1] = value
            ws.append(new_row)
        else:
            # Update existing row
            for col_name, value in FIELD_MAP.items():
                idx = col_idx(col_name)
                if idx:
                    ws.cell(row=target_row, column=idx, value=value)
            for col_name, value in checklist_values.items():
                idx = col_idx(col_name)
                if idx:
                    ws.cell(row=target_row, column=idx, value=value)

        atomic_save(wb, EXCEL_FILE)

    return jsonify({"ok": True})


# Federal contracting lifecycle (Phase 3G). "Active Bid" is a PURSUE status — the
# only stage that authorizes folder-creating automation (the watcher creates the
# folder on its next poll).
_CLOSED_STATES = {"Closed Won", "Closed Lost", "Closed Cancelled", "Closed Withdrawn"}
_WORKFLOW_STAGES = ({"Recommended", "Manual Review", "Prospect", "Active Bid",
                     "Submitted", "Awarded", "Archived"} | _CLOSED_STATES)

# Server-enforced valid transitions (Phase 3G-A). Anything not listed → 409.
# Unknown/legacy source statuses are permissive (migration safety — existing
# Active Bid / Awarded / legacy rows are never blocked).
_TRANSITIONS = {
    "Recommended":      {"Prospect", "Archived"},
    "Manual Review":    {"Prospect", "Recommended", "Archived"},
    "Prospect":         {"Active Bid", "Recommended", "Archived"},
    "Active Bid":       {"Submitted", "Archived"},
    "Submitted":        {"Awarded", "Closed Lost", "Closed Cancelled", "Closed Withdrawn", "Archived"},
    "Awarded":          {"Closed Won", "Archived"},
    "Archived":         {"Recommended"},
    # Closed → reopen to Submitted (admin/API path only; audited BID_REOPENED).
    "Closed Won":       {"Submitted"},
    "Closed Lost":      {"Submitted"},
    "Closed Cancelled": {"Submitted"},
    "Closed Withdrawn": {"Submitted"},
}

# Lifecycle columns written on the relevant transitions (Phase 3G-H).
_LIFECYCLE_COLS = ["Submitted Date", "Submitted By", "Submission Status", "Submission Method",
                   "Closed Date", "Closed Reason", "Result Status"]
_RESULT_OF = {"Closed Won": "Won", "Closed Lost": "Lost",
              "Closed Cancelled": "Cancelled", "Closed Withdrawn": "Withdrawn"}


@app.route("/api/bids/status", methods=["POST"])
def set_bid_status():
    """Move a bid to a new workflow stage. Writes ONLY the Status column (never
    clobbers other fields) and logs an automation_audit event. No folder is
    created here — that stays gated to the watcher on a PURSUE status."""
    import datetime as _dt
    body = request.get_json(force=True) or {}
    bid_id = str(body.get("id", "")).strip()
    title  = str(body.get("title", "")).strip()
    stage  = str(body.get("status", "")).strip()
    if stage not in _WORKFLOW_STAGES:
        return jsonify({"error": f"invalid stage '{stage}'"}), 400
    if not EXCEL_FILE.exists():
        return jsonify({"error": "Excel file not found"}), 500

    with workbook_lock(EXCEL_FILE):
        wb = openpyxl.load_workbook(EXCEL_FILE, data_only=False)
        ws = wb["Bid Tracker"]
        raw_headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]

        def col_idx(name):
            for i, h in enumerate(raw_headers):
                if h.replace('\xa0', '').strip().lower() == name.lower():
                    return i + 1
            return None

        # Ensure lifecycle columns exist so we can persist submitted/closed fields.
        for h in _LIFECYCLE_COLS:
            if col_idx(h) is None:
                ws.cell(row=1, column=ws.max_column + 1, value=h)
                raw_headers.append(h)

        id_col, title_col = col_idx("Solicitation / Notice ID"), col_idx("Bid / Opportunity Name")
        status_col, updated_col = col_idx("Status"), col_idx("Last Updated")
        is_auto_id = bid_id.startswith("bt-")

        target_row = None
        if not is_auto_id and bid_id and id_col:
            for row in ws.iter_rows(min_row=2):
                if str(row[id_col - 1].value or "").strip() == bid_id:
                    target_row = row[0].row
                    break
        if target_row is None and title_col and title:
            for row in ws.iter_rows(min_row=2):
                if str(row[title_col - 1].value or "").strip().lower() == title.lower():
                    target_row = row[0].row
                    break
        if target_row is None:
            return jsonify({"error": "bid not found"}), 404

        old_status = str(ws.cell(row=target_row, column=status_col).value or "").strip() if status_col else ""
        # Server-enforced transition validation (Phase 3G-A). Known source states
        # must follow the lifecycle; unknown/legacy sources are permissive.
        if old_status in _TRANSITIONS and stage != old_status and stage not in _TRANSITIONS[old_status]:
            return jsonify({"error": f"invalid transition '{old_status}' → '{stage}'",
                            "allowed": sorted(_TRANSITIONS[old_status])}), 409

        now_dt = _dt.datetime.now().strftime("%Y-%m-%d %H:%M")
        today  = _dt.date.today().isoformat()
        if status_col:
            ws.cell(row=target_row, column=status_col, value=stage)
        if updated_col:
            ws.cell(row=target_row, column=updated_col, value=now_dt)

        def _set(col_name, value, overwrite=True):
            c = col_idx(col_name)
            if c and (overwrite or not str(ws.cell(target_row, c).value or "").strip()):
                ws.cell(target_row, c, value=value)

        # Lifecycle field writes (Phase 3G-E / 3G-H).
        if stage == "Submitted" and old_status == "Active Bid":
            _set("Submitted Date", today, overwrite=False)   # keep original submit date
            _set("Submitted By", str(body.get("submitted_by") or "bid-tracker-ui"), overwrite=False)
            _set("Submission Status", "Submitted")
            if body.get("submission_method"):
                _set("Submission Method", str(body.get("submission_method")))
        elif stage in _CLOSED_STATES:
            _set("Closed Date", today)
            _set("Result Status", _RESULT_OF[stage])
            _set("Closed Reason", str(body.get("closed_reason") or _RESULT_OF[stage]))
        elif stage == "Submitted" and old_status in _CLOSED_STATES:
            _set("Closed Date", "")            # reopened — clear closed fields
            _set("Result Status", "")
            _set("Closed Reason", "")

        atomic_save(wb, EXCEL_FILE)

    # "Bid On This" (→ Active Bid) is the sanctioned, user-initiated folder
    # trigger. Create the estimate folder NOW (idempotent) instead of waiting for
    # the 60s watcher. create_estimate_folders takes its OWN workbook lock, so we
    # call it AFTER releasing ours above to avoid a nested/deadlocked lock.
    folder_created, folder_path, err = False, "", ""
    if stage == "Active Bid":
        try:
            import json as _json
            import create_estimate_folders as cef
            cef.run()   # creates folders for Active-Bid rows not yet in the manifest
            manifest = _json.loads(cef.MANIFEST.read_text()) if cef.MANIFEST.exists() else {}
            folder_path = (manifest.get(bid_id) or {}).get("folder", "")
            folder_created = bool(folder_path)
        except Exception as e:
            err = str(e)

    # Forensic audit trail for the transition (create_estimate_folders logs the
    # folder-creation event itself; this records the user's stage change).
    # Audit action by transition (Phase 3G-I).
    if stage == "Submitted" and old_status in _CLOSED_STATES:
        _action = "BID_REOPENED"
    elif stage == "Submitted":
        _action = "BID_SUBMITTED"
    elif stage in _CLOSED_STATES:
        _action = "BID_CLOSED"
    elif stage == "Awarded":
        _action = "PROMOTION_STATUS_CHANGED"
    else:
        _action = "set_workflow_status"
    try:
        from capture_engine import log_automation
        log_automation(opportunity_id=bid_id, title=title, user="bid-tracker-ui",
                       trigger=f"{old_status or 'none'}→{stage}", action=_action,
                       result=("error" if err else "ok"),
                       folder_created=folder_created, folder_path=folder_path,
                       error_message=err)
    except Exception:
        pass
    return jsonify({"ok": True, "status": stage, "oldStatus": old_status,
                    "folderCreated": folder_created, "folderPath": folder_path,
                    "error": err or None})


# Phase 3H — per-bid metadata (follow-up, award probability, government response).
# Additive: written by /api/bids/meta ONLY; never touches Status / the transition
# engine. Normalized key → Excel column.
_META_MAP = {
    "awardProbability":   "Award Probability",
    "followUpStatus":     "Follow-Up Status",
    "lastFollowUp":       "Last Follow-Up Date",
    "nextFollowUp":       "Next Follow-Up Date",
    "followUpNotes":      "Follow-Up Notes",
    "coName":             "CO Name",
    "coEmail":            "CO Email",
    "awardDecisionDate":  "Award Decision Date",
    "debriefRequested":   "Debrief Requested",
    "debriefReceived":    "Debrief Received",
}


@app.route("/api/bids/meta", methods=["POST"])
def update_bid_meta():
    """Write additive per-bid metadata (follow-up / award probability / government
    response). Does NOT modify Status or the transition engine — only the meta
    columns in _META_MAP. Body: {id, title, fields: {normalizedKey: value, ...}}."""
    body = request.get_json(force=True) or {}
    bid_id = str(body.get("id", "")).strip()
    title  = str(body.get("title", "")).strip()
    fields = body.get("fields") or {}
    if not isinstance(fields, dict) or not fields:
        return jsonify({"error": "no fields provided"}), 400
    if not EXCEL_FILE.exists():
        return jsonify({"error": "Excel file not found"}), 500

    with workbook_lock(EXCEL_FILE):
        wb = openpyxl.load_workbook(EXCEL_FILE, data_only=False)
        ws = wb["Bid Tracker"]
        raw = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]

        def ci(name):
            for i, h in enumerate(raw):
                if h.replace('\xa0', '').strip().lower() == name.lower():
                    return i + 1
            return None

        # Ensure only the meta columns being written exist.
        to_write = {}
        for key, value in fields.items():
            col_name = _META_MAP.get(key)
            if not col_name:
                continue
            if ci(col_name) is None:
                ws.cell(row=1, column=ws.max_column + 1, value=col_name)
                raw.append(col_name)
            to_write[col_name] = value

        id_col, title_col = ci("Solicitation / Notice ID"), ci("Bid / Opportunity Name")
        target_row = None
        if bid_id and not bid_id.startswith("bt-") and id_col:
            for row in ws.iter_rows(min_row=2):
                if str(row[id_col - 1].value or "").strip() == bid_id:
                    target_row = row[0].row; break
        if target_row is None and title_col and title:
            for row in ws.iter_rows(min_row=2):
                if str(row[title_col - 1].value or "").strip().lower() == title.lower():
                    target_row = row[0].row; break
        if target_row is None:
            return jsonify({"error": "bid not found"}), 404

        for col_name, value in to_write.items():
            c = ci(col_name)
            if c:
                ws.cell(target_row, c, value=("" if value is None else value))
        atomic_save(wb, EXCEL_FILE)

    return jsonify({"ok": True, "updated": list(to_write.keys())})


@app.route("/api/automation-audit")
def automation_audit():
    """Recent automation_audit events (folder creation, status changes, etc.)."""
    try:
        from capture_engine import read_recent
        return jsonify(read_recent(int(request.args.get("limit", 100))))
    except Exception as e:
        return jsonify({"error": str(e), "events": []})


# Financial Hub service layer — the ONLY path Bid Tracker uses to create projects.
# Bid Tracker never writes accounting/project/payroll tables directly (Section 7).
HUB_URL = "http://localhost:8000"
_PROMO_COLS = ["Financial Hub Project ID", "Job Number", "Promotion Date", "Promotion Status"]


@app.route("/api/bids/promote", methods=["POST"])
def promote_bid_to_hub():
    """Promote an AWARDED capture bid to a Financial Hub project via the Hub
    service layer, then write the returned IDs back to the workbook. Gates on
    Awarded status server-side and never overwrites an existing promotion."""
    import datetime as _dt
    import requests as _rq
    body = request.get_json(force=True) or {}
    bid_id = str(body.get("id", "")).strip()
    title  = str(body.get("title", "")).strip()
    if not EXCEL_FILE.exists():
        return jsonify({"error": "Excel file not found"}), 500

    # ── 1. Read row, gate on Awarded, dedup, gather payload ───────────────────
    with workbook_lock(EXCEL_FILE):
        wb = openpyxl.load_workbook(EXCEL_FILE, data_only=False)
        ws = wb["Bid Tracker"]

        def headers():
            return [str(c.value).strip() if c.value is not None else "" for c in ws[1]]

        def col_idx(name):
            for i, h in enumerate(headers()):
                if h.replace('\xa0', '').strip().lower() == name.lower():
                    return i + 1
            return None

        # Ensure promotion columns exist (lazy, additive).
        for h in _PROMO_COLS:
            if col_idx(h) is None:
                ws.cell(row=1, column=ws.max_column + 1, value=h)

        id_col, title_col = col_idx("Solicitation / Notice ID"), col_idx("Bid / Opportunity Name")
        target_row = None
        if bid_id and not bid_id.startswith("bt-") and id_col:
            for row in ws.iter_rows(min_row=2):
                if str(row[id_col - 1].value or "").strip() == bid_id:
                    target_row = row[0].row; break
        if target_row is None and title_col and title:
            for row in ws.iter_rows(min_row=2):
                if str(row[title_col - 1].value or "").strip().lower() == title.lower():
                    target_row = row[0].row; break
        if target_row is None:
            return jsonify({"error": "bid not found"}), 404

        def cell(name):
            c = col_idx(name)
            return str(ws.cell(target_row, c).value or "").strip() if c else ""

        status = cell("Status")
        if status != "Awarded":
            return jsonify({"error": f"Only Awarded bids can be promoted (current: '{status or 'none'}')"}), 409

        # Never re-promote (Section 4 — never overwrite).
        if cell("Financial Hub Project ID"):
            return jsonify({"success": True, "already": True,
                            "project_id": cell("Financial Hub Project ID"),
                            "job_number": cell("Job Number")}), 200

        def _num(s):
            try: return float(str(s).replace(",", "").replace("$", "")) or 0.0
            except (TypeError, ValueError): return 0.0

        payload = {
            "bid_id":        bid_id or (cell("Solicitation / Notice ID") or title),
            "title":         title or cell("Bid / Opportunity Name"),
            "agency":        cell("Agency / Department"),
            "location":      ", ".join(x for x in (cell("City"), cell("State")) if x),
            "naics":         cell("NAICS"),
            "award_amount":  _num(cell("Award Amount")) or _num(cell("Bid Amount")),
            "award_date":    _dt.date.today().isoformat(),
            "performed_by":  "bid-tracker-ui",
        }

    # ── 2. Call the Financial Hub service layer (lock released) ───────────────
    try:
        resp = _rq.post(f"{HUB_URL}/api/projects/promote-capture-bid", json=payload, timeout=30)
    except Exception as e:
        return jsonify({"error": f"Financial Hub unreachable: {e}"}), 502

    if resp.status_code not in (200, 409):
        return jsonify({"error": "Hub promotion failed", "detail": resp.text[:400]}), 502
    data = resp.json() if resp.content else {}
    # 409 = already promoted in Hub → recover the IDs and write them back.
    project_id = str(data.get("project_id") or data.get("financial_hub_id") or "")
    job_number = str(data.get("job_number") or "")
    if not project_id or not job_number:
        return jsonify({"error": "Hub returned no IDs", "detail": data}), 502

    # ── 3. Write IDs back to the workbook (never overwrite existing) ──────────
    promoted_at = _dt.datetime.now().strftime("%Y-%m-%d %H:%M")
    with workbook_lock(EXCEL_FILE):
        wb = openpyxl.load_workbook(EXCEL_FILE, data_only=False)
        ws = wb["Bid Tracker"]
        raw = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]

        def ci(name):
            for i, h in enumerate(raw):
                if h.replace('\xa0', '').strip().lower() == name.lower():
                    return i + 1
            return None

        # Ensure promotion columns exist (persisted by atomic_save below).
        for h in _PROMO_COLS:
            if ci(h) is None:
                ws.cell(row=1, column=ws.max_column + 1, value=h)
                raw.append(h)   # keep raw in sync so ci() resolves the new column

        id_col, title_col = ci("Solicitation / Notice ID"), ci("Bid / Opportunity Name")
        trow = None
        if bid_id and not bid_id.startswith("bt-") and id_col:
            for row in ws.iter_rows(min_row=2):
                if str(row[id_col - 1].value or "").strip() == bid_id:
                    trow = row[0].row; break
        if trow is None and title_col and title:
            for row in ws.iter_rows(min_row=2):
                if str(row[title_col - 1].value or "").strip().lower() == title.lower():
                    trow = row[0].row; break
        if trow is not None:
            writes = {"Financial Hub Project ID": project_id, "Job Number": job_number,
                      "Promotion Date": promoted_at, "Promotion Status": "Promoted"}
            for name, val in writes.items():
                c = ci(name)
                if c and not str(ws.cell(trow, c).value or "").strip():   # never overwrite
                    ws.cell(trow, c, value=val)
        atomic_save(wb, EXCEL_FILE)

    # ── 4. Audit (capture side) ───────────────────────────────────────────────
    try:
        from capture_engine import log_automation
        log_automation(opportunity_id=payload["bid_id"], title=payload["title"],
                       user="bid-tracker-ui", trigger="Awarded→Project Created",
                       action="PROJECT_PROMOTED", result="ok",
                       folder_path=f"FH project {project_id} / {job_number}")
    except Exception:
        pass

    return jsonify({"success": True, "project_id": project_id, "job_number": job_number,
                    "financial_hub_id": project_id})


@app.route("/api/projects/<path:job_number>/summary")
def project_summary(job_number):
    """Read-only proxy to the Financial Hub job-cost dashboard for a promoted
    project. Surfaces procurement (committed cost + PO count) and cost/margin so
    the Bid Tracker can show project financials — reads only, no DB writes."""
    import requests as _rq
    try:
        r = _rq.get(f"{HUB_URL}/api/projects/db/{job_number}/dashboard", timeout=15)
    except Exception as e:
        return jsonify({"error": f"Financial Hub unreachable: {e}"}), 502
    if r.status_code == 404:
        return jsonify({"error": "project not found in Financial Hub"}), 404
    if r.status_code != 200:
        return jsonify({"error": "Financial Hub error", "detail": r.text[:300]}), 502
    data = r.json()
    # Phase 3F — merge the per-category job-cost breakdown (Budget/Actual/Remaining).
    try:
        rc = _rq.get(f"{HUB_URL}/api/projects/db/{job_number}/cost-summary", timeout=15)
        if rc.status_code == 200:
            cs = rc.json()
            data["categories"]  = cs.get("categories", [])
            data["cost_totals"] = cs.get("totals", {})
    except Exception:
        pass
    return jsonify(data)


@app.route("/api/bids/initialize-project", methods=["POST"])
def bid_initialize_project():
    """Phase 4A — generate the Project Startup Pack for a promoted bid via the Hub
    service layer (idempotent). Bid Tracker never writes project tables directly."""
    import requests as _rq
    body = request.get_json(force=True) or {}
    job_number = str(body.get("jobNumber") or body.get("job_number") or "").strip()
    if not job_number:
        return jsonify({"error": "jobNumber required — promote the bid to a project first"}), 400
    try:
        r = _rq.post(f"{HUB_URL}/api/projects/{job_number}/startup",
                     json={"performed_by": "bid-tracker-ui"}, timeout=60)
    except Exception as e:
        return jsonify({"error": f"Financial Hub unreachable: {e}"}), 502
    if r.status_code != 200:
        return jsonify({"error": "startup failed", "detail": r.text[:300]}), 502
    return jsonify(r.json())


@app.route("/api/projects/<path:job_number>/startup")
def project_startup(job_number):
    """Read-only proxy of a project's startup-pack status."""
    import requests as _rq
    try:
        r = _rq.get(f"{HUB_URL}/api/projects/{job_number}/startup", timeout=15)
    except Exception as e:
        return jsonify({"error": f"Financial Hub unreachable: {e}"}), 502
    return jsonify(r.json()), (200 if r.status_code == 200 else r.status_code)


@app.route("/api/health")
def health():
    return jsonify({
        "status": "ok",
        "excel":  str(EXCEL_FILE),
        "exists": EXCEL_FILE.exists(),
        "hits_tab": _latest_hits_tab(),
    })


# ── System health + job ledger — restored from the legacy backend, which had
# these but never made it into this file when it forked. The live UI (System
# Health widget, Job Ledger tab) has been calling these and getting 404s. ──────
try:
    from system_health import build_system_health as _build_system_health
    _SYS_HEALTH_OK = True
except ImportError:
    _SYS_HEALTH_OK = False

@app.route("/api/system/health")
def system_health():
    if not _SYS_HEALTH_OK:
        return jsonify({"error": "system_health module unavailable"}), 503
    return jsonify(_build_system_health())


try:
    from job_manager import get_jobs as _get_jobs, get_job as _get_job, \
                           get_summary as _get_summary, reset_job as _reset_job
    _JOBS_OK = True
except ImportError:
    _JOBS_OK = False

def _jobs_unavailable():
    return jsonify({"error": "job_manager unavailable"}), 503

@app.route("/ops/jobs/summary")
def ops_summary():
    if not _JOBS_OK:
        return jsonify({"failed_today": 0, "queued": 0, "running": 0, "success_24h": 0, "by_flow": {}})
    return jsonify(_get_summary())

@app.route("/ops/jobs")
def ops_jobs():
    if not _JOBS_OK:
        return jsonify([])
    return jsonify(_get_jobs(
        flow=request.args.get("flow"),
        status=request.args.get("status"),
        limit=int(request.args.get("limit", 50)),
    ))

@app.route("/ops/jobs/<job_id>")
def ops_job(job_id):
    if not _JOBS_OK:
        return _jobs_unavailable()
    job = _get_job(job_id)
    return jsonify(job) if job else (jsonify({"error": "not found"}), 404)

@app.route("/ops/jobs/<job_id>/retry", methods=["POST"])
def ops_retry(job_id):
    if not _JOBS_OK:
        return _jobs_unavailable()
    _reset_job(job_id)
    return jsonify({"ok": True})


if __name__ == "__main__":
    print(f"  API server → http://localhost:5050")
    print(f"  Excel file → {EXCEL_FILE}")
    print(f"  File exists: {EXCEL_FILE.exists()}")
    # M-17: bind to localhost only. The API is unauthenticated; 0.0.0.0 exposed
    # it to the whole LAN. The Bid Tracker UI reaches it via the vite proxy
    # (target 127.0.0.1:5050), so localhost binding does not affect the UI.
    app.run(host="127.0.0.1", port=5050, debug=False)
